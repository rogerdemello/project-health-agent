"""Robust Excel parser with fuzzy column detection."""

from datetime import datetime, date
from difflib import get_close_matches

import openpyxl
from loguru import logger

from app.models import ParsedProject, Task, Milestone, Comment
from app.config import COLUMN_ALIASES
from app.utils import parse_date, parse_variance, parse_percent


def _fuzzy_match(header: str, aliases: dict[str, list[str]]) -> str | None:
    """Match a column header to a canonical key using tiered matching."""
    hl = header.strip().lower()

    # Tier 1: exact canonical match
    if hl in aliases:
        return hl

    # Tier 2: exact alias match
    for canonical, candidates in aliases.items():
        for alias in candidates:
            if hl == alias:
                return canonical

    # Tier 3: alias contains header or header contains alias (for short headers)
    for canonical, candidates in aliases.items():
        for alias in candidates:
            if len(alias) > 4 and (alias in hl or hl in alias):
                return canonical

    # Tier 4: fuzzy fallback with strict threshold
    flat_aliases = list({a for candidates in aliases.values() for a in candidates if len(a) > 4})
    matches = get_close_matches(hl, flat_aliases, n=1, cutoff=0.75)
    if matches:
        best = matches[0]
        for canonical, candidates in aliases.items():
            if best in candidates:
                return canonical

    return None


def _find_sheet(wb: openpyxl.Workbook) -> str | None:
    """Find the main data sheet (not Comments or Summary)."""
    exclude = {"comments", "summary"}
    for sn in wb.sheetnames:
        if sn.strip().lower() not in exclude:
            return sn
    return None


def parse_project(filepath: str) -> ParsedProject:
    """Parse an Excel project plan into a structured ParsedProject model."""
    logger.info(f"Parsing project file: {filepath}")
    wb = openpyxl.load_workbook(filepath, data_only=True)
    project = ParsedProject()

    # ---- Summary sheet ----
    if "Summary" in wb.sheetnames:
        ws = wb["Summary"]
        summary: dict[str, str] = {}
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
            if row[0] and row[1] is not None:
                key = str(row[0]).strip()
                val = str(row[1]).strip() if not isinstance(row[1], (datetime, date)) else str(row[1])
                summary[key] = val
        project.raw_summary = summary
        project.name = summary.get("Project Name", "")
        project.manager = summary.get("Project Manager", "")
        project.completion_percent = parse_percent(summary.get("% Complete"))
        project.status = summary.get("Project Status", "")
        today_val = summary.get("Today's Date")
        if today_val:
            parsed = parse_date(today_val)
            if parsed:
                project.assessment_date = parsed

    # ---- Comments sheet ----
    if "Comments" in wb.sheetnames:
        ws = wb["Comments"]
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
            text_parts = [str(c).strip() for c in row if c and isinstance(c, str) and len(str(c).strip()) > 3]
            if text_parts:
                project.comments.append(Comment(
                    author=row[2] if len(row) > 2 and row[2] else "",
                    date=str(row[3]) if len(row) > 3 and row[3] else "",
                    text=" ".join(text_parts),
                ))

    # ---- Main data sheet ----
    sheet_name = _find_sheet(wb)
    if not sheet_name:
        logger.warning("No data sheet found; returning empty project.")
        return project

    ws = wb[sheet_name]
    headers = [str(c.value).strip() if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]

    # Build column mapping
    col_map: dict[str, int] = {}
    for i, h in enumerate(headers, 1):
        if h:
            matched = _fuzzy_match(h, COLUMN_ALIASES)
            if matched:
                col_map[matched] = i
                logger.debug(f"  Mapped '{h}' -> '{matched}' (col {i})")

    logger.info(f"Detected columns: {list(col_map.keys())}")

    # Detect project name from first task if not in summary
    task_name_col = col_map.get("task_name")
    if not project.name and task_name_col:
        for row in ws.iter_rows(min_row=2, max_row=min(4, ws.max_row), values_only=True):
            if len(row) >= task_name_col and row[task_name_col - 1]:
                val = str(row[task_name_col - 1]).strip()
                if val and val not in ("#UNPARSEABLE", "", "?"):
                    project.name = val
                    break

    # ---- Parse tasks ----
    task_id = 0
    phase_col = col_map.get("phase")
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        task_id += 1
        get_col = lambda name: row[col_map[name] - 1] if col_map.get(name) and len(row) >= col_map[name] else None

        task_name = str(get_col("task_name") or "").strip()
        if not task_name or task_name == "#UNPARSEABLE":
            continue

        start = parse_date(get_col("start_date"))
        end = parse_date(get_col("end_date"))
        bl_start = parse_date(get_col("baseline_start"))
        bl_finish = parse_date(get_col("baseline_finish"))

        raw_var = get_col("variance")
        variance = parse_variance(raw_var) if raw_var is not None else None
        if variance is None and bl_finish and end:
            variance = (bl_finish - end).days

        schedule_health = str(get_col("schedule_health") or "").strip()

        critical_raw = get_col("critical")
        critical = False
        if critical_raw is not None:
            if isinstance(critical_raw, bool):
                critical = critical_raw
            else:
                critical = str(critical_raw).strip().lower() in ("yes", "true", "1", "y")

        at_risk_raw = get_col("at_risk")
        at_risk = False
        if at_risk_raw is not None:
            at_risk = str(at_risk_raw).strip().lower() in ("yes", "high", "true", "1", "y")

        on_hold_raw = get_col("on_hold")
        on_hold = False
        if on_hold_raw is not None:
            on_hold = str(on_hold_raw).strip().lower() in ("yes", "true", "1", "y")

        phase_name = str(get_col("phase") or "").strip() if phase_col else ""
        comment_text = str(get_col("comments") or "").strip()
        owner = str(get_col("owner") or "").strip()
        pct = parse_percent(get_col("percent_complete"))
        status = str(get_col("status") or "").strip()

        task = Task(
            id=task_id,
            name=task_name,
            owner=owner,
            status=status,
            percent_complete=pct,
            start_date=start,
            end_date=end,
            baseline_start=bl_start,
            baseline_finish=bl_finish,
            variance_days=variance,
            critical=critical,
            schedule_health=schedule_health,
            phase=phase_name,
            at_risk=at_risk,
            on_hold=on_hold,
            comments=comment_text,
        )
        project.tasks.append(task)

        # Detect milestones from phase-level tasks or completed phases
        if task.phase and task.phase != task.name:
            project.milestones.append(Milestone(
                name=task.name,
                due_date=bl_finish or end,
                actual_date=end,
                completed=status.lower() in ("completed", "done"),
                percent_complete=pct,
                variance_days=variance,
            ))

    # Also infer milestones from level/ancestors pattern: tasks with phase name != task name are milestones
    if not project.milestones:
        phase_names: set[str] = set()
        for t in project.tasks:
            if t.phase and t.phase != t.name and t.phase not in ("#UNPARSEABLE", "", "?"):
                phase_names.add(t.phase)
        for pn in sorted(phase_names):
            phase_tasks = [t for t in project.tasks if t.phase == pn and t.name != pn]
            if phase_tasks:
                first = phase_tasks[0]
                project.milestones.append(Milestone(
                    name=pn,
                    due_date=first.baseline_finish or first.end_date,
                    actual_date=first.end_date,
                    completed=all(t.status.lower() in ("completed", "done") for t in phase_tasks),
                    percent_complete=sum(t.percent_complete for t in phase_tasks) / len(phase_tasks),
                    variance_days=first.variance_days,
                ))

    logger.info(f"Parsed {len(project.tasks)} tasks, {len(project.milestones)} milestones, {len(project.comments)} comments")
    return project
